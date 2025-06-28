import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
from datetime import datetime

# 📌 페이지 설정
st.set_page_config(
    page_title="중량물 작업계획서",
    page_icon="📝",
    layout="centered"
)

st.title("📝 중량물 작업계획서 자동 생성기")
st.markdown("작업 정보를 입력하고 **엑셀 파일 생성** 버튼을 누르세요.")

# ✅ 입력 필드 정의
fields = {
    "부서명": "",
    "작업일자": datetime.today().strftime('%Y-%m-%d'),
    "작업 장소": "",
    "장비 No.": "",
    "직종": "",
    "성명": "",
    "작업 내용": "",
    "근무시간": "",
    "건강상태": "",
    "비고": ""
}

data = {}

# ✅ 입력 폼
with st.form("work_plan_form"):
    for key, default in fields.items():
        data[key] = st.text_input(key, value=default)
    submitted = st.form_submit_button("📁 엑셀 파일 생성")

# ✅ 엑셀 파일 생성
if submitted:
    wb = Workbook()
    ws = wb.active
    ws.title = "중량물 작업계획서"

    # 서식 지정
    title_font = Font(size=14, bold=True)
    header_font = Font(size=12, bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_wrap = Alignment(wrap_text=True)

    # 제목 행
    ws.merge_cells("A1:B1")
    ws["A1"] = "중량물 작업계획서"
    ws["A1"].font = title_font
    ws["A1"].alignment = align_center

    # 내용 행
    row = 3
    for key, value in data.items():
        ws[f"A{row}"] = key
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = header_font
        ws[f"A{row}"].alignment = align_center
        ws[f"B{row}"].alignment = align_wrap
        row += 1

    # 열 너비 조정
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50

    # 메모리 버퍼에 저장
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # 다운로드 버튼
    st.success("✅ 엑셀 파일이 성공적으로 생성되었습니다.")
    st.download_button(
        label="📥 엑셀 파일 다운로드",
        data=buffer,
        file_name=f"중량물작업계획서_{data['작업일자']}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
